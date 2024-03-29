# importing libraries & modules
import os
import re
import cv2
import time
import config
import random
import platform
import pyqrcode
import numpy as np
import tkinter as tk
try:
    import frontend_api as fi
except:
    import App.frontend_api as fi
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl import Workbook, load_workbook


# check OS to set GUI size
def chkos():
    """
    This function identifies the Operating system being used to
    set the geometry of the Windows in this app.
    """

    oname = platform.system()
    if oname == "Windows":
        config.screen1geo = "430x300"
        config.screen1_5geo = "490x145"
        config.screen2geo = "500x310"
        config.screen3geo = "360x125"
        config.screen4geo = "585x458"
        config.screen5geo = "690x470"
        config.screen6geo = "420x190"
        config.screen7geo = "300x250"

    elif oname == "Linux":
        config.screen1geo = "375x300"
        config.screen1_5geo = "390x145"
        config.screen2geo = "520x310"
        config.screen3geo = "320x125"
        config.screen4geo = "680x458"
        config.screen5geo = "770x460"
        config.screen6geo = "495x190"
        config.screen7geo = "360x270"

    else:
        config.screen1geo = "375x300"
        config.screen1_5geo = "390x145"
        config.screen2geo = "520x310"
        config.screen3geo = "320x125"
        config.screen4geo = "250x258"
        config.screen5geo = "760x460"
        config.screen6geo = "455x190"
        config.screen7geo = "300x300"


# code for QR scanner
def scanner():
    """
    This function is a module that manages the aspects of a QR Scanner
    using various sub modules.

    :return: it returns the data obtained from the QR Code.
    """

    # start device camera
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)
    cap.set(4, 480)
    time.sleep(2)

    # find content of QR
    def decode(im: object) -> object:
        """
        This function decodes image given by camera to isolate QR code present in it.

        :param im: it is the image obtained from the device camera.
        :return: it returns the data obtained by decoding the image.
        """

        decodedObjects = pyzbar.decode(im)
        return decodedObjects

    # font for hull
    font = cv2.FONT_HERSHEY_SIMPLEX

    # scan the QR
    def app():
        """
        This function uses the device camera to scan the QR shown to it &
        helps save the QR if user wants to.

        :return: it returns the data obtained by the scanner.
        """

        decodedObject = ""
        while cap.isOpened():
            ret, frame = cap.read()
            im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decodedObjects = decode(im)
            for decodedObject in decodedObjects:
                points = decodedObject.polygon
                if len(points) > 4:
                    hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                    hull = list(map(tuple, np.squeeze(hull)))
                else:
                    hull = points
                n = len(hull)
                for j in range(0, n):
                    # while scanning
                    cv2.line(frame, hull[j], hull[(j+1) % n], (255, 0, 0), 3)
                    x = decodedObject.rect.left
                    y = decodedObject.rect.top
                    cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0, 255, 255), 2, cv2.LINE_AA)

            try:
                barCode = str(decodedObject.data)
                bar = barCode
            except:
                bar = ""

            cv2.imshow('frame', frame)
            key = cv2.waitKey(1)
            # binding q key as shortcut to close camera
            if key & 0xFF == ord('q') or key == 27:
                cap.release()
                cv2.destroyAllWindows()
                break
            # binding s key as shortcut to save image in front of camera
            elif key & 0xFF == ord('s'):
                if (bar != "") or (len(bar) != 0):
                    iname = "./scan/" + bar + ".png"
                else:
                    iname = "./scan/" + str(random.randint(1, 101)) + ".png"
                cv2.imwrite(iname, frame)
                messagebox.showinfo("INFO", "QR Saved")

        # check if any QR detected
        try:
            barCode = str(decodedObject.data)
            bar = barCode
            return bar
        except:
            messagebox.showerror("ALERT", "No QR Detected")
            return 0

    # start scanner
    bar = app()
    # close camera
    cap.release()
    cv2.destroyAllWindows()
    return bar


# QR Scanner code
def QRScan():
    """
    This function is a module that manages aspects of participant entry
    using serveral sub modules.
    """

    # call scanner
    def callscan():
        """
        This function calls the QR Scanner to scan the QR of participants.
        """

        resp = scanner()
        if resp == 0:
            screen7.focus_force()
        else:
            qrpid.set(resp)
            screen7.focus_force()

    # entry marker
    def marker():
        """
        This function validates the QR provided by a participant and marks their entry
        by matching it against data in the server.
        """

        try:
            if (qrpid.get() != "") and (qreve.get() != ""):
                uid = qrpid.get()[2:-1]
                eve = qreve.get()
                resp = fi.mark_entry(p_id=uid, event=eve)
                if resp == 0:
                    messagebox.showerror("ALERT", "No participant registered with this QR ID in this event")
                elif resp == 1:
                    messagebox.showinfo("Success", "Participant entry marked. \nEntry Granted")
                else:
                    messagebox.showerror("ALERT", "Participant already entered. \nEntry Denied")
                screen7.focus_force()
                qrpid.set("")
            elif qrpid.get() != "":
                messagebox.showerror("ALERT", "No Event selected")
            elif qreve.get() != "":
                messagebox.showerror("ALERT", "No QR ID found")
            else:
                messagebox.showerror("ALERT", "Fields Incomplete")
        except:
            messagebox.showerror("ALERT", "Unable to connect to the server")
        finally:
            screen7.focus_force()

    # GUI of entry marker
    screen7 = Toplevel(config.screen4)
    screen7.title("Event Entry Management")
    screen7.geometry(config.screen7geo)
    screen7.resizable(False, False)
    screen7.config(background=colr)
    screen7.focus_force()
    icon = PhotoImage(file="./resc/qr-code-scan.png")
    screen7.iconphoto(False, icon)
    try:
        evts = fi.get_events()
        qrpid = StringVar()
        qreve = StringVar()
        label = Label(screen7, text="Participant Entry", bg=colr, font=("Times New Roman", 20, 'bold'))
        label.configure(foreground="white", anchor="center")
        label.grid(row=1, column=1, padx=20, pady=(20, 15), columnspan=3)
        label = Label(screen7, width=15, text="Select Event : ", bg=colr)
        label.configure(foreground="white")
        label.grid(row=3, column=1, padx=5, pady=10)
        screen7.entry1 = ttk.Combobox(screen7, width=17, textvariable=qreve, state="readonly")
        screen7.entry1.grid(row=3, column=2, padx=5, pady=10, columnspan=1)
        screen7.entry1['values'] = evts
        screen7.entry1.current()
        label = Label(screen7, width=15, text="QR ID : ", bg=colr)
        label.configure(foreground="white")
        label.grid(row=4, column=1, padx=5, pady=10)
        screen7.entryname = Entry(screen7, width=20, textvariable=qrpid)
        screen7.entryname.grid(row=4, column=2, padx=5, pady=10, columnspan=2)
        buton = Button(screen7, width=15, text="Scan (ctrl + s)", command=callscan)
        buton.grid(row=5, column=1, padx=15, pady=(20, 0), columnspan=1)
        # binding Ctrl + s key as shortcut to open scanner
        screen7.bind("<Control-s>", lambda event=None: buton.invoke())
        button = Button(screen7, width=15, text="Mark (↵)", command=marker)
        button.grid(row=5, column=2, padx=15, pady=(20, 0), columnspan=1)
        # binding Enter key as shortcut to mark the entry
        screen7.bind('<Return>', lambda event=None: button.invoke())
        label = Label(screen7, width=35, text="Press 'Esc' or 'q' to close camera \nPress 's' to save image.", bg=colr)
        label.configure(foreground="#767777")
        label.grid(row=6, column=1, padx=5, pady=10, columnspan=2)
    except:
        messagebox.showerror("ALERT", "Unable to connect to the server")
    finally:
        screen7.focus_force()

    # set focus to management window on closing
    config.screen4.focus_force()


# code to register & generate QR for participant
def QRP():
    """
    This function is a module that manages aspects of participant registration
    using serveral sub modules.
    """

    # call scanner
    def callscan():
        """
        This function calls the QR Scanner to register existing participant in another event.
        """

        resp = scanner()
        if resp == 0:
            config.screen5.focus_force()
        else:
            qrID.set(resp)
            config.screen5.focus_force()

    # code for GUI of QR generator
    def QRGen():
        """
        This function loads the GUI of participant registration window.
        """

        gcolor = "#161a2d"
        config.screen5 = Toplevel(config.screen4)
        config.screen5.title("QR Generator")
        config.screen5.geometry(config.screen5geo)
        config.screen5.resizable(False, False)
        config.screen5.config(background=gcolor)
        icon = PhotoImage(file="./resc/laptop.png")
        config.screen5.iconphoto(False, icon)
        try:
            evts = fi.get_events()
            label = Label(config.screen5, text="Event Registration", bg=gcolor, font=("Times New Roman", 20, 'bold'))
            label.configure(foreground="white", anchor="center")
            label.grid(row=0, column=2, padx=5, pady=5, columnspan=4)
            label = Label(config.screen5, text="Enter all details or QR-ID of participant", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=1, column=1, padx=5, pady=10, columnspan=3)
            label = Label(config.screen5, text="Enter Name : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=2, column=1, padx=5, pady=10)
            config.screen5.entryname = Entry(config.screen5, width=30, textvariable=qrName)
            config.screen5.entryname.grid(row=2, column=2, padx=5, pady=10, columnspan=2)
            config.screen5.entryname.focus_set()
            label = Label(config.screen5, text="Enter Phno : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=3, column=1, padx=5, pady=10)
            config.screen5.entryphno = Entry(config.screen5, width=30, textvariable=qrphno)
            config.screen5.entryphno.grid(row=3, column=2, padx=5, pady=10, columnspan=2)
            label = Label(config.screen5, text="Enter Email : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=4, column=1, padx=5, pady=10)
            config.screen5.entrymail = Entry(config.screen5, width=30, textvariable=qrmail)
            config.screen5.entrymail.grid(row=4, column=2, padx=5, pady=10, columnspan=2)
            label = Label(config.screen5, text="QR ID : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=5, column=1, padx=5, pady=10)
            config.screen5.entry = Entry(config.screen5, width=15, textvariable=qrID)
            config.screen5.entry.grid(row=5, column=2, padx=10, pady=10, columnspan=1, sticky='w')
            sbtn = Button(config.screen5, width=10, text="Scan (ctrl + s)", command=callscan)
            sbtn.grid(row=5, column=3, padx=5, pady=10, sticky='e')
            # binding Ctrl + s key as shortcut to open scanner
            config.screen5.bind("<Control-s>", lambda event=None: sbtn.invoke())
            ttk.Separator(config.screen5, orient=HORIZONTAL).grid(column=1, row=6, columnspan=3, sticky='ew')
            label = Label(config.screen5, text="1st Event Name : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=7, column=1, padx=5, pady=10)
            label = Label(config.screen5, text="2nd Event Name : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=8, column=1, padx=5, pady=10)
            config.screen5.entry1 = ttk.Combobox(config.screen5, width=27, textvariable=qrevent1, state="readonly")
            config.screen5.entry1.grid(row=7, column=2, padx=5, pady=10, columnspan=2)
            config.screen5.entry1['values'] = evts
            config.screen5.entry1.current()
            config.screen5.entry2 = ttk.Combobox(config.screen5, width=27, textvariable=qrevent2, state="readonly")
            config.screen5.entry2.grid(row=8, column=2, padx=5, pady=10, columnspan=2)
            config.screen5.entry2['values'] = evts
            config.screen5.entry2.current()
            label = Label(config.screen5, text="QR Code : ", bg=gcolor)
            label.configure(foreground="white")
            label.grid(row=9, column=1, padx=5, pady=10)
            button = Button(config.screen5, width=10, text="Generate (↵)", command=QRCodeGenerate)
            button.grid(row=9, column=2, padx=5, pady=10, columnspan=1)
            # binding Enter key as shortcut to generate QR
            config.screen5.bind('<Return>', lambda event=None: button.invoke())
            buton = Button(config.screen5, width=10, text="Clear (ctrl + r)", command=QRClear)
            buton.grid(row=9, column=3, padx=5, pady=10, columnspan=1)
            # binding Ctrl + r key as shortcut to clear fields
            config.screen5.bind("<Control-r>", lambda event=None: buton.invoke())
            config.screen5.imageLabel = Label(config.screen5, background=gcolor)
            config.screen5.imageLabel.grid(row=2, column=4, rowspan=9, columnspan=3, padx=(10, 5), pady=10)
            image = Image.open("./resc/wait.png")
            image = image.resize((350, 350), Image.ANTIALIAS)
            image = ImageTk.PhotoImage(image)
            config.screen5.imageLabel.config(image=image)
            config.screen5.imageLabel.photo = image

        except:
            messagebox.showerror("ALERT", "Unable to connect to the server")

        finally:
            config.screen5.focus_force()

    # reload wait image
    def ld():
        """
        This function loads the waiting screen image after fields are cleared.
        """

        image = Image.open("./resc/wait.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        config.screen5.imageLabel.config(image=image)
        config.screen5.imageLabel.photo = image

    # code for clearing values of GUI fields
    def QRClear():
        """
        This function clears the fields in participant registration window.
        """

        config.screen5.entryname.focus_set()
        qrName.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        qrevent2.set("")
        qrID.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        config.screen5.imageLabel.config(image=image)
        config.screen5.imageLabel.photo = image
        config.screen5.after(500, ld)

    # code to generate QR with participant data
    def QRCodeGenerate():
        """
        This function generates QR Code for participant after validating the information provided by them.
        """

        if (qrName.get() != '') and (qrphno.get() != '') and (qrmail.get() != '') and (qrevent1.get() != ''):
            if len(qrphno.get()) == 10:
                try:
                    rege = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                    if re.search(rege, qrmail.get()):
                        content = qrName.get().lower() + "-" + qrphno.get()
                        i = QRdatamgSQL(content)
                        if i == 0:
                            messagebox.showerror("ALERT", "Internal error in registration")
                        elif i == 2:
                            messagebox.showinfo("Pay Attention", "Participant already registered in event 1. \nRegistration for event 2 complete")
                        elif i == 3:
                            messagebox.showinfo("Pay Attention", "Participant already registered in event 2. \nRegistration for event 1 complete")
                        elif i == 4:
                            messagebox.showerror("ALERT", "Participant already registered in provided events. \nRegistration Aborted")
                        elif i == 1:
                            qrGenerate = pyqrcode.create(content)
                            qrCodePath = './data/'
                            qrCodeName = qrCodePath + qrphno.get() + ".png"
                            qrGenerate.png(qrCodeName, scale=10)
                            image = Image.open(qrCodeName)
                            image = image.resize((350, 350), Image.ANTIALIAS)
                            image = ImageTk.PhotoImage(image)
                            config.screen5.imageLabel.config(image=image)
                            config.screen5.imageLabel.photo = image
                            QRdatamgXL()
                        config.screen5.focus_force()
                    else:
                        messagebox.showerror("ALERT", "Invalid Email ID")
                        config.screen5.focus_force()
                        config.screen5.entrymail.focus_set()
                except:
                    messagebox.showerror("ALERT", "Error in Generating QR or Phone NaN")
                    config.screen5.focus_force()
                    config.screen5.entryphno.focus_set()
            else:
                messagebox.showerror("ALERT", "Invalid Phone Number")
                config.screen5.focus_force()
                config.screen5.entryphno.focus_set()
        elif (qrID.get() != '') and (qrevent1.get() != ''):
            autofil()
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            config.screen5.focus_force()

    # code for autofill
    def autofil():
        """
        This function registers an existing participant in another event.
        """

        uid = qrID.get()[2:-1]
        eve = [qrevent1.get(), qrevent2.get()]
        if qrevent2.get() == "":
            eve = [qrevent1.get()]
        try:
            resp = fi.add_part(p_id=uid, name="", email_id="", phone="", events=eve)
            if resp == 0:
                messagebox.showerror("ALERT", "No Registration found on this QR ID \nRegistration Aborted")
            elif resp == 2:
                messagebox.showinfo("Pay Attention", "Participant already registered in event 1. \nRegistration for event 2 complete")
            elif resp == 3:
                messagebox.showinfo("Pay Attention", "Participant already registered in event 2. \nRegistration for event 1 complete")
            elif resp == 4:
                messagebox.showerror("ALERT", "Participant already registered in provided events. \nRegistration Aborted")
            else:
                messagebox.showinfo("Success", "Registration Completed")
        except:
            messagebox.showerror("ALERT", "Unable to connect to the server")
        finally:
            config.screen5.focus_force()

    # code to add participant data to excel sheet
    def QRdatamgXL():
        """
        This function adds the participant data to an Excel sheet in local machine
        as soon as the participant data is added to the server.
        """

        # path for excel file
        path = "./data/regdata.xlsx"
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            wb = Workbook(path)
            wb.save(path)
        except PermissionError:
            messagebox.showerror("Alert", "File access denied. \nClose the excel sheet to fix this issue.")
        wb = load_workbook(path)
        sheet = wb.active
        row = (qrName.get(), qrphno.get(), qrmail.get(), qrevent1.get(), qrevent2.get())
        sheet.append(row)
        wb.save(path)

    # code to add participant data to SQL
    def QRdatamgSQL(content: str) -> int:
        """
        This function contacts the server to add participant data in the database.

        :param content: it contains the data to be used as participant ID from the QR Code.
        :return: it returns the response provided by the server, to check if data has been stored.
        """

        eve = [qrevent1.get(), qrevent2.get()]
        if qrevent2.get() == "":
            eve = [qrevent1.get()]
        try:
            resp = fi.add_part(p_id=content, name=qrName.get(), email_id=qrmail.get(), phone=qrphno.get(), events=eve)
            return resp
        except:
            messagebox.showerror("ALERT", "Unable to connect to the server")
        finally:
            config.screen5.focus_force()

    qrName = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()
    qrID = StringVar()
    QRGen()


# code to maintain event list
def eventmgm():
    """
    This function is a module that loads the GUI of event management window &
    contains sub modules to manage aspects of events.
    """

    # clear fields
    def clrevent():
        """
        This function clears all the fields in the event management window.
        """

        adevent.focus_set()
        evename.set("")
        evedate.set("")
        evetime.set("")

    # clear placeholder of date field
    def clrdt(event: object):
        """
        This function is used to monitor widget selection to
        clear the placeholder value.

        :param event: checks mouse interrupt.
        """

        evedate.set("")

    # clear placeholder of time field
    def clrti(event: object):
        """
        This function is used to monitor widget selection to
        clear the placeholder value.

        :param event: checks mouse interrupt.
        """

        evetime.set("")

    # add events to database
    def addevent():
        """
        This function contacts the server to add an event data & performs validation
        on the data provided & stops if event already exists.
        """

        if (evename.get() != "") and (evedate.get() != "") and (evetime.get() != ""):
            try:
                resp = fi.add_event(name=evename.get(), date=evedate.get(), time=evetime.get())
                if resp == 0:
                    messagebox.showerror("ALERT", "Another event entry with same name already exists")
                    adevent.focus_set()
                elif resp == 2:
                    messagebox.showerror("ALERT", "Wrong date format: \nCorrect format: YYYY-MM-DD")
                    adevedt.focus_set()
                elif resp == 3:
                    messagebox.showerror("ALERT", "Wrong Time format: \nCorrect format: HH:MM")
                    adeveti.focus_set()
                else:
                    messagebox.showinfo("Success", "Event added successfully")
            except:
                messagebox.showerror("ALERT", "Unable to connect to the server")
            finally:
                screen6.focus_force()
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            screen6.focus_force()

    # remove events from database
    def remevent():
        """
        This function contacts the server to remove an event from the database by
        validating the information & stops the process if information is invalid
        or if any participant is registered to the event.
        """

        if (evename.get() != "") and (evedate.get() != "") and (evetime.get() != ""):
            try:
                resp = fi.remove_event(name=evename.get(), date=evedate.get(), time=evetime.get())
                if resp == 0:
                    messagebox.showerror("ALERT", "Someone is registered in this event")
                    adevent.focus_set()
                elif resp == 1:
                    messagebox.showinfo("Success", "Event removed successfully")
                elif resp == 2:
                    messagebox.showerror("ALERT", "Wrong date format: \nCorrect format: YYYY-MM-DD")
                    adevedt.focus_set()
                elif resp == 3:
                    messagebox.showerror("ALERT", "Wrong Time format: \nCorrect format: HH:MM")
                    adeveti.focus_set()
                elif resp == 4:
                    messagebox.showerror("ALERT", "Invalid event details")
            except:
                messagebox.showerror("ALERT", "Unable to connect to the server")
            finally:
                screen6.focus_force()
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            screen6.focus_force()

    # GUI code for event manager
    screen6 = Toplevel(config.screen4)
    screen6.title("Event Manager")
    screen6.geometry(config.screen6geo)
    screen6.resizable(False, False)
    screen6.config(background="green")
    icon = PhotoImage(file="./resc/team-management.png")
    screen6.iconphoto(False, icon)
    screen6.focus_force()
    evename = StringVar()
    evetime = StringVar()
    evedate = StringVar()
    label = Label(screen6, text="Event Management", bg="green", font=("Times New Roman", 20, 'bold'))
    label.configure(foreground="white", anchor="center")
    label.grid(row=0, column=1, padx=(17, 0), pady=(10, 15), columnspan=4)
    lbl = Label(screen6, text="Event Name", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=1, column=1, padx=(40, 5), pady=5, columnspan=1)
    adevent = Entry(screen6, width='17', textvariable=evename)
    adevent.grid(row=1, column=2, padx=5, pady=5, columnspan=1)
    lbl = Label(screen6, text="Event Date", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=2, column=1, padx=(40, 5), pady=5, columnspan=1)
    adevedt = Entry(screen6, width='17', textvariable=evedate)
    adevedt.insert(0, 'YYYY-MM-DD')
    adevedt.bind("<FocusIn>", clrdt)
    adevedt.grid(row=2, column=2, padx=5, pady=5, columnspan=1)
    lbl = Label(screen6, text="Event Time", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=3, column=1, padx=(40, 5), pady=5, columnspan=1)
    adeveti = Entry(screen6, width='17', textvariable=evetime)
    adeveti.insert(0, 'HH:MM')
    adeveti.bind("<FocusIn>", clrti)
    adeveti.grid(row=3, column=2, padx=5, pady=5, columnspan=1)
    bnt = Button(screen6, text="Clear (ctrl + r)", command=clrevent, width=18)
    bnt.grid(row=1, column=3, padx=5, pady=5, columnspan=2)
    nbt = Button(screen6, text="Add Event (ctrl + a)", command=addevent, width=18)
    nbt.grid(row=2, column=3, padx=5, pady=5, columnspan=2)
    tnb = Button(screen6, text="Remove Event (ctrl + d)", command=remevent, width=18)
    tnb.grid(row=3, column=3, padx=5, pady=5, columnspan=2)
    adevent.focus_set()
    # binding Ctrl + a key as shortcut to add event
    screen6.bind("<Control-a>", lambda event=None: nbt.invoke())
    # binding Ctrl + d key as shortcut to remove event
    screen6.bind("<Control-d>", lambda event=None: tnb.invoke())
    # binding Ctrl + r key as shortcut to clear fields
    screen6.bind("<Control-r>", lambda event=None: bnt.invoke())

    # code to monitor event management window close event
    def on_closing():
        """
        This function monitors event management window close event &
        reopens tasks window.
        """

        config.screen4.deiconify()
        screen6.destroy()
    screen6.protocol("WM_DELETE_WINDOW", on_closing)


# code to generate report
def report_gen():
    """
    This function is a module that contacts the server to generate report of
    participants & their attendance in various events.
    """

    try:
        rep = fi.get_report()
        # path to report excel file
        p = "./data/report.xlsx"
        try:
            wb = load_workbook(p)
        except FileNotFoundError:
            wb = Workbook(p)
            wb.save(p)
        try:
            wb = load_workbook(p)
            sheet = wb.active
            sheet.delete_cols(1, 20)
            sheet.delete_rows(1, 1000)
            row = ("S.No.", "QR ID", "Name", "E-mail", "Phone no.")
            sheet.append(row)
            wb.save(p)
            count = 1
            for i in rep:
                row = (count, i[0], i[1], i[2], i[3])
                col = 6
                e_count = 1
                count += 1
                sheet.append(row)
                for x in i[4].split(","):
                    sheet.cell(row=1, column=col).value = "Event " + str(e_count)
                    sheet.cell(row=count, column=col).value = x[:-1]
                    sheet.cell(row=1, column=col + 1).value = "E-" + str(e_count) + " Entry"
                    sheet.cell(row=count, column=col + 1).value = x[-1].replace("1", "Not Entered").replace("2", "Entered")
                    col += 2
                    e_count += 1
                wb.save(p)
            messagebox.showinfo("Success", "Report Generated successfully \nPath: " + p)
        except PermissionError:
            messagebox.showerror("Alert", "File access denied. \nClose the excel sheet OR Run program as Administrator to fix this issue.")
    except:
        messagebox.showerror("ALERT", "Unable to connect to the server")
    finally:
        # set focus to management window on closing
        config.screen4.focus_force()


# code to manage user tasks
def mgm_page():
    """
    This function loads the GUI of main tasks window of the app & guides its users of its purpose,
    It contains the participant registration, participant entry, event management & report generation modules.
    """

    # GUI for organizer management
    config.screen3.withdraw()
    config.screen4 = config.screen4
    config.screen4 = Toplevel(config.screen3)
    config.screen4.title("Select")
    config.screen4.geometry(config.screen4geo)
    config.screen4.resizable(False, False)
    config.screen4.config(background=colr)
    icon = PhotoImage(file="./resc/process.png")
    config.screen4.iconphoto(False, icon)
    config.screen4.focus_force()
    label = Label(config.screen4, text="Event Registration & Verification Using QR", bg=colr, fg="white", font=("Times New Roman", 20, 'bold'))
    label.grid(row=1, column=1, padx=5, pady=(20, 30), columnspan=3)
    label = Label(config.screen4, text="Participant Registration", bg=colr, fg="white", font=("Times New Roman", 12, 'bold'))
    label.grid(row=2, column=1, padx=(30, 40), pady=15, columnspan=1)
    btn = Button(config.screen4, width=15, borderwidth=0, text="Registry (ctrl + g)", command=QRP)
    btn.grid(row=3, column=1, padx=(30, 40), pady=10, columnspan=1)
    label = Label(config.screen4, text="Register participants in one or more \nevents & Generate QR code, \nunique for everyone", bg=colr, fg="white")
    label.grid(row=4, column=1, padx=(30, 40), pady=10, columnspan=1)
    label = Label(config.screen4, text="Participant Verification", bg=colr, fg="white", font=("Times New Roman", 12, 'bold'))
    label.grid(row=2, column=3, padx=50, pady=15, columnspan=1)
    bnt = Button(config.screen4, width=15, borderwidth=0, text="Entry (ctrl + s)", command=QRScan)
    bnt.grid(row=3, column=3, padx=50, pady=10, columnspan=1)
    label = Label(config.screen4, text="Verify and mark participant's entry, \nusing the QR code provided, \nfor each event", bg=colr, fg="white")
    label.grid(row=4, column=3, padx=50, pady=10, columnspan=1)
    ttk.Separator(config.screen4, orient=HORIZONTAL).grid(column=1, row=5, columnspan=3, sticky='ew')
    ttk.Separator(config.screen4, orient=HORIZONTAL).grid(column=2, row=2, rowspan=9, sticky='ns')
    label = Label(config.screen4, text="Event Management", bg=colr, fg="white", font=("Times New Roman", 12, 'bold'))
    label.grid(row=6, column=1, padx=(30, 40), pady=15, columnspan=1)
    tbn = Button(config.screen4, width=15, borderwidth=0, text="Manage (ctrl + e)", command=eventmgm)
    tbn.grid(row=7, column=1, padx=(30, 40), pady=10, columnspan=1)
    label = Label(config.screen4, text="Add and remove events to be organized, \nalong with their date and time", bg=colr, fg="white")
    label.grid(row=8, column=1, padx=(30, 40), pady=10, columnspan=1)
    label = Label(config.screen4, text="Report Generator", bg=colr, fg="white", font=("Times New Roman", 12, 'bold'))
    label.grid(row=6, column=3, padx=50, pady=15, columnspan=1)
    ttbn = Button(config.screen4, width=15, borderwidth=0, text="Report (ctrl + r)", command=report_gen)
    ttbn.grid(row=7, column=3, padx=5, pady=10, columnspan=1)
    label = Label(config.screen4, text="Generate report for all events and \nparticipants along with their details", bg=colr, fg="white")
    label.grid(row=8, column=3, padx=50, pady=10, columnspan=1)
    # binding Ctrl + g key as shortcut to open participant adding window
    config.screen4.bind("<Control-g>", lambda event=None: btn.invoke())
    # binding Ctrl + s key as shortcut to open participant entry window
    config.screen4.bind("<Control-s>", lambda event=None: bnt.invoke())
    # binding Ctrl + e key as shortcut to open event management window
    config.screen4.bind("<Control-e>", lambda event=None: tbn.invoke())
    # binding Ctrl + r key as shortcut to generate report
    config.screen4.bind("<Control-r>", lambda event=None: ttbn.invoke())

    # code to monitor main tasks window / app close event
    def on_closing():
        """
        This function monitors the close event main tasks window of the app.
        it also terminates the entire app.
        """

        config.screen1.destroy()
    config.screen4.protocol("WM_DELETE_WINDOW", on_closing)


# GUI & code for login & signup
def main_page():
    """
    This function is a module that loads the GUI of login window and contains
    user login & registration modules, which are managed by sub modules.
    """

    # code to clear login data fields after successful login
    def clrlogin():
        """
        This function clears the fields of login window
        """

        username_verify.set("")
        password_verify.set("")
        mgm_page()

    # code to organizer management
    def register_user():
        """
        This function manages user registration success window.
        """

        # code to monitor screen1 close event
        def on_closing():
            """
            This function monitors the Registration window close event.
            """

            config.screen2.destroy()
            config.screen1.deiconify()
        config.screen2.protocol("WM_DELETE_WINDOW", on_closing)

        # GUI for user add success
        def disab():
            """
            This function sets the GUI of user registration success window.
            """

            config.screen1_5 = Toplevel(config.screen1)
            config.screen1_5.title("Success")
            config.screen1_5.geometry(config.screen1_5geo)
            config.screen1_5.resizable(False, False)
            config.screen1_5.config(background="green")
            config.screen1_5.focus_force()

            # code to call login success screen
            def calllog():
                """
                This function monitors the user Registration window close event &
                clears the fields present in it & reopens the admin login window.
                """

                username.set("")
                emailid.set("")
                phno.set("")
                password.set("")
                config.screen3.deiconify()
                config.screen1_5.destroy()
                config.screen2.destroy()
                adminlogin()

            label = Label(config.screen1_5, text="", bg="green")
            label.grid(row=1, column=1)
            label = Label(config.screen1_5, text="Registeration Success", width='30', bg="green", font=("Times New Roman", 20, 'bold'))
            label.configure(foreground="white")
            label.grid(pady=5, row=2, column=1, columnspan=1)
            bttn = Button(config.screen1_5, text="OK (↵)", width="15", command=calllog)
            bttn.grid(pady=5, row=3, column=1, columnspan=1)
            # binding Enter key as shortcut to proceed
            config.screen1_5.bind('<Return>', lambda event=None: bttn.invoke())

        disab()

    # registry data validation
    def valinp():
        """
        This function validates the data provided by admin to add more users,
        and sends the validated data to the server.
        """

        if (username.get() != "") and (emailid.get() != "") and (phno.get() != "") and (password.get() != "") and (config.perm_entry.get() != "Select"):
            if len(phno.get()) == 10:
                rege = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if re.search(rege, emailid.get()):
                    if re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{8,}', password.get()):
                        perm = config.perm_entry.get()
                        perm = 2 if perm == "Admin" else 1
                        try:
                            resp = fi.add_user(name=username.get(), email_id=emailid.get(), password=password.get(), phone=int(phno.get()), perm=perm)
                            if resp == 0:
                                messagebox.showerror("ALERT", "User email already exists")
                                config.screen2.focus_force()
                            else:
                                register_user()
                        except:
                            messagebox.showerror("ALERT", "Unable to connect to the server")
                        finally:
                            config.screen2.focus_force()
                    else:
                        messagebox.showerror("ALERT", "Password not Strong")
                        config.screen2.focus_force()
                else:
                    messagebox.showerror("ALERT", "Invalid Email")
                    config.screen2.focus_force()
            else:
                messagebox.showerror("ALERT", "Invalid Phone Number")
                config.screen2.focus_force()
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            config.screen2.focus_force()

    # GUI code for adding user
    def register():
        """
        This function sets the GUI of user registration window.
        """

        config.screen3.withdraw()
        config.screen2 = Toplevel(config.screen1)
        config.screen2.title("Register")
        config.screen2.geometry(config.screen2geo)
        config.screen2.resizable(False, False)
        config.screen2.config(background=colr)
        icon = PhotoImage(file="./resc/add.png")
        config.screen2.iconphoto(False, icon)
        config.screen2.focus_force()
        labl = Label(config.screen2, text="Please enter user information", width="30", bg=colr)
        labl.configure(foreground="white", font=("Times New Roman", 20, 'bold'))
        labl.grid(row=1, column=1, padx=5, pady=5, columnspan=2)
        labl = Label(config.screen2, text="User Name", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry = Entry(config.screen2, textvariable=username)
        username_entry.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry.focus_set()
        labl = Label(config.screen2, text="Email ID", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=2, column=2, padx=5, pady=5, columnspan=1)
        emailid_entry = Entry(config.screen2, textvariable=emailid)
        emailid_entry.grid(row=3, column=2, padx=5, pady=5, columnspan=1)
        labl = Label(config.screen2, text="Phone Number", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        phno_entry = Entry(config.screen2, textvariable=phno)
        phno_entry.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(config.screen2, text="Password", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=4, column=2, padx=5, pady=5, columnspan=1)
        password_entry = Entry(config.screen2, show="*", textvariable=password)
        password_entry.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        labl = Label(config.screen2, text="", width="30", bg=colr)
        labl.grid(row=6, column=1, padx=5, pady=5, columnspan=2)
        labl = Label(config.screen2, text="Permission : ", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        config.perm_entry = ttk.Combobox(config.screen2, textvariable=rights, width="17", values=["Select", "Admin", "User"], state="readonly")
        config.perm_entry.current(0)
        config.perm_entry.grid(row=7, column=2, columnspan=1, pady=5)
        labl = Label(config.screen2, text="", bg=colr)
        labl.grid(row=8, column=1, columnspan=2)
        regbtn = Button(config.screen2, text="Sumbit (↵)", width='18', command=valinp)
        regbtn.grid(row=9, column=1, padx=5, pady=5, columnspan=2)
        # binding Enter key as shortcut to proceed
        config.screen2.bind('<Return>', lambda event=None: regbtn.invoke())

        # code to monitor screen2 close event
        def on_closing():
            """
            This function monitors the user registration window close event &
            it reopens the admin login success window.
            """

            config.screen3.deiconify()
            config.screen2.destroy()
        config.screen2.protocol("WM_DELETE_WINDOW", on_closing)

    # GUI if user is admin
    def adminlogin():
        """
        This functions sets the GUI if user is an admin.
        """

        config.screen3.geometry(config.screen3geo)
        label = Label(config.screen3, text="Login Success", width='30', bg="green")
        label.configure(foreground="white", font=("Times New Roman", 16, 'bold'))
        label.grid(row=1, column=1, pady=5, columnspan=1)
        bttnn = Button(config.screen3, text="OK (↵)", width="15", command=clrlogin)
        bttnn.grid(row=2, column=1, pady=5, columnspan=1)
        bttn = Button(config.screen3, text="Add User (ctrl + a)", width="15", command=register)
        bttn.grid(row=3, column=1, pady=5, columnspan=1)
        # binding Enter key as shortcut to proceed
        config.screen3.bind('<Return>', lambda event=None: bttnn.invoke())
        # binding Ctrl + a key as shortcut to access user adding screen
        config.screen3.bind("<Control-a>", lambda event=None: bttn.invoke())

    # GUI if user is user
    def userlogin():
        """
        This function sets the GUI if the user is an organizer.
        """

        config.screen3.geometry(config.screen3geo)
        label = Label(config.screen3, text="", bg="green")
        label.grid(row=1, column=1, pady=5)
        label = Label(config.screen3, text="Login Success", width='30', bg="green")
        label.configure(foreground="white", font=("Times New Roman", 16, 'bold'))
        label.grid(row=2, column=1, pady=5)
        bttn = Button(config.screen3, text="OK (↵)", width="10", command=clrlogin)
        bttn.grid(row=3, column=1, pady=5)
        # binding Enter key as shortcut to proceed
        config.screen3.bind('<Return>', lambda event=None: bttn.invoke())

    # code for GUI & user details verification
    def login_verify():
        """
        This function sets the GUI for login success screen &
        it sends the login credentials to be verified & authenticated to the server.
        """

        config.screen1.withdraw()
        config.screen3 = Toplevel(config.screen1)
        config.screen3.title("Info")
        config.screen3.geometry(config.screen3geo)
        config.screen3.resizable(False, False)
        config.screen3.config(background="green")
        config.screen3.focus_force()
        icon = PhotoImage(file="./resc/check.png")
        config.screen3.iconphoto(False, icon)

        # code to monitor screen3 close event
        def on_closing():
            """
            This function monitors login success screen close event,
            and reopens login window.
            It is the equivalent of logging out.
            """

            clrlogin()
            config.screen1.deiconify()
            config.screen3.destroy()
        config.screen3.protocol("WM_DELETE_WINDOW", on_closing)

        try:
            resp = fi.login(uid=username_verify.get(), password=password_verify.get())
            if resp == 2:
                adminlogin()
            elif resp == 1:
                userlogin()
            elif resp == 0:
                on_closing()
                messagebox.showerror("ALERT", "Invalid User/password")
                username_entry1.focus_set()
            else:
                on_closing()
                messagebox.showerror("ALERT", "Invalid User")
                username_entry1.focus_set()
        except:
            on_closing()
            messagebox.showerror("ALERT", "No Internet")
            username_entry1.focus_set()

    # check if fields are complete
    def chk_login_verify():
        """
        This function ensures that the fields in login screen are completed.
        """

        if (username_verify.get() != "") and (password_verify.get() != ""):
            login_verify()
        elif username_verify.get() == "":
            messagebox.showerror("ALERT", "Username Field Incomplete")
            username_entry1.focus_set()
        elif password_verify.get() == "":
            messagebox.showerror("ALERT", "Password Field Incomplete")
            password_entry1.focus_set()
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            username_entry1.focus_set()

    # code for login GUI
    config.screen1 = Tk()
    config.screen1.title("Login")
    config.screen1.geometry(config.screen1geo)
    config.screen1.config(background=colr)
    config.screen1.resizable(False, False)
    icon = PhotoImage(file="./resc/login.png")
    config.screen1.iconphoto(False, icon)
    username = StringVar()
    password = StringVar()
    emailid = StringVar()
    phno = StringVar()
    rights = StringVar()
    username_verify = StringVar()
    password_verify = StringVar()
    label = Label(text="", bg=colr)
    label.grid(row=1, column=1)
    label = Label(text="Please Enter your Login \nInformation", width='30', bg=colr)
    label.configure(foreground="white", font=("Times New Roman", 18, 'bold'))
    label.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
    label = Label(text="User Email ID : ", width='30', bg=colr)
    label.configure(foreground="white")
    label.grid(row=4, column=1, padx=25, pady=5, columnspan=1)
    username_entry1 = Entry(width="21", textvariable=username_verify)
    username_entry1.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
    username_entry1.focus_set()
    label = Label(text="Password : ", width='30', bg=colr)
    label.configure(foreground="white")
    label.grid(row=6, column=1, padx=5, pady=5, columnspan=1)
    password_entry1 = Entry(width='21', show="*", textvariable=password_verify)
    password_entry1.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
    label = Label(text="", bg=colr)
    label.grid(row=8, column=1)
    btnn = Button(text="Login (↵)", width="18", command=chk_login_verify)
    btnn.grid(row=9, column=1, padx=5, pady=5, columnspan=1)
    # binding Enter key as shortcut to login
    config.screen1.bind('<Return>', lambda event=None: btnn.invoke())

    # monitor app close
    def on_closing(event: object):
        """
        This function is used to monitor app close event,
        it binds the escape to it.

        :param event: checks keyboard interrupt.
        """

        sys.exit()
    # binding Escape key as shortcut to close app
    config.screen1.bind('<Escape>', on_closing)
    config.screen1.mainloop()


# securing code from import exploits
if __name__ == "__main__":
    # setting a main color theme
    colr = "#1c44a5"

    # checking OS to set GUI geometry
    chkos()

    # start program by calling login module
    main_page()
