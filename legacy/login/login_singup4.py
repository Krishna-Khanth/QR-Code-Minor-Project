import os
import re 
import cv2
import time
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles.colors import *
#from tkinter import messagebox

def mgm_page():
    path = "./data/regdata.xlsx"               
    def QRScan():
        cap = cv2.VideoCapture(0)
        cap.set(3,640)
        cap.set(4,480)
        time.sleep(2)

        def decode(im): 
            decodedObjects = pyzbar.decode(im)
            return decodedObjects

        font = cv2.FONT_HERSHEY_SIMPLEX

        def regbk(bar):
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    bar = bar.replace("'","")
                    value = cell.value + "'"
                    if(cell.value == bar):
                        print(cell.coordinate)
                        cell.fill = PatternFill(bgColor="00FF00", fill_type="solid")
                        cell.font = Font(color="00FF00")
                        wb.save(path)

        def regdb():
            print ("Sharad")
            """
            iss function se connect kar existing SQL db se,
            aur check kar if data being scanned db mei hai ya nahi,
            if present add present in present column
            and scanner wale hull ka color green
            if already marked present hull color red,
            GUI alert pop karva.
            """

        def app():
            decodedObject = ""
            while(cap.isOpened()):
                ret, frame = cap.read()
                im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                decodedObjects = decode(im)
                for decodedObject in decodedObjects: 
                    points = decodedObject.polygon
                    if len(points) > 4 : 
                      hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                      hull = list(map(tuple, np.squeeze(hull)))
                    else : 
                      hull = points;         
                    n = len(hull)     
                    for j in range(0,n):
                      cv2.line(frame, hull[j], hull[ (j+1) % n], (255,0,0), 3)
                    x = decodedObject.rect.left
                    y = decodedObject.rect.top
                    cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0,255,255), 2, cv2.LINE_AA)
                    
                cv2.imshow('frame', frame)
                key = cv2.waitKey(1)
                if key & 0xFF == ord('q'):
                    break
                elif key & 0xFF == ord('s'):
                    cv2.imwrite('./scan/capt.png', frame)     
                
            barCode = str(decodedObject.data)            
            barC = barCode.split('-')
            bar = barC[1]
            regbk(bar)
            regdb()
                                         
        app()
        cap.release()
        cv2.destroyAllWindows() 
        
    def QRGen():
        root.title("QR Generator")
        root.config(background="green")
        """
        btn.pack_forget()
        bnt.pack_forget()
        """
        root.geometry("425x615")
        label = Label(text="Enter Name : ", bg="green")
        label.grid(row=0, column=1, padx=5, pady=5)
        root.entry = Entry(width=30, textvariable=qrInput)
        root.entry.grid(row=0, column=2, padx=5, pady=5, columnspan=2)
        label = Label(text="Enter Phno : ", bg="green")
        label.grid(row=1, column=1, padx=5, pady=5)
        root.entry = Entry(width=30, textvariable=qrphno)
        root.entry.grid(row=1, column=2, padx=5, pady=5, columnspan=2)
        label = Label(text="Enter Email : ", bg="green")
        label.grid(row=2, column=1, padx=5, pady=5)
        root.entry = Entry(width=30, textvariable=qrmail)
        root.entry.grid(row=2, column=2, padx=5, pady=5, columnspan=2)    
        label = Label(text="Event Name : ", bg="green")
        label.grid(row=3, column=1, padx=5, pady=5)    
        root.entry = Entry(width=30, textvariable=qrevent1)
        root.entry.grid(row=3, column=2, padx=5, pady=5, columnspan=2)
        label = Label(text="Event Name : ", bg="green")
        label.grid(row=4, column=1, padx=5, pady=5)    
        root.entry = Entry(width=30, textvariable=qrevent2)
        root.entry.grid(row=4, column=2, padx=5, pady=5, columnspan=2)
        label = Label(text="QR Code : ", bg="green")
        label.grid(row=5, column=1, padx=5, pady=5)
        button = Button(width=10, text="Generate", command=QRCodeGenerate)
        button.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        buton = Button(width=10, text="Clear", command=QRClear)
        buton.grid(row=5, column=3, padx=5, pady=5, columnspan=1)
        label = Label(text="QR Code : ", bg="green")
        label.grid(row=5, column=1, padx=5, pady=5)
        root.imageLabel = Label(root, background="green")
        root.imageLabel.grid(row=6, column=1, columnspan=3, padx=5, pady=5)
        image = Image.open("./resc/wait.jpg")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        root.imageLabel.config(image=image)
        root.imageLabel.photo = image
        button = Button(width=13, text="Registration Open", command=QROpen)
        button.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        buton = Button(width=13, text="Back", command=GenScan)
        buton.grid(row=7, column=2, padx=5, pady=5, columnspan=1)
        buton = Button(width=13, text="Registration Closed", command=QRClose)
        buton.grid(row=7, column=3, padx=5, pady=5, columnspan=1)
        
    def QRClear():        
        qrInput.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        root.imageLabel.config(image=image)
        root.imageLabel.photo = image
        
    def QROpen():
        print("OPEN")
        """
        reactivate all disabled fields & buttons
        """
        
    def QRClose():
        print("Close")        
        """
        ask for CAPTCHA type verification, then disable 
        all fields and buttons except Registration Open,
        break databases according to event names
        """
        
    def QRCodeGenerate():    
        qrString = qrInput.get()
        qrPhno = qrphno.get()
        qrMail = qrmail.get()
        qrevent1 = qrevent1.get()
        if (qrString != '') and (qrPhno != '') and (qrMail != '') and (qrevent1 != ''):
            if (len(qrPhno) == 10):
                regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if(re.search(regex,qrMail)):  
                    content = qrString + "-" + qrPhno
                    qrGenerate = pyqrcode.create(content)
                    qrCodePath = './data/'
                    qrCodeName = qrCodePath + qrPhno + ".png"
                    qrGenerate.png(qrCodeName, scale=10)
                    image = Image.open(qrCodeName)
                    image = image.resize((350, 350), Image.ANTIALIAS)
                    image = ImageTk.PhotoImage(image)
                    root.imageLabel.config(image=image)
                    root.imageLabel.photo = image
                    QRdatamgSQL()
                    QRdatamgXL()
                else:  
                    messagebox.showerror("ERROR", "Invalid Email ID")
            else:
                messagebox.showerror("ERROR", "Invalid Phone Number")
        else:
            messagebox.showerror("ERROR", "Fields Incomplete")

    def QRdatamgXL():
        wb = load_workbook(path)
        sheet = wb.active
        row = ((qrInput.get(), qrphno.get(), qrmail.get(), qrevent1.get(), qrevent2.get()))
        sheet.append(row)
        wb.save(path)

    def QRdatamgSQL():
        print("SHARAD")
        """
        ek new function bana to check if data 
        being entered is already in db.
        ek new function bana for SQL db creation,
        iss function se SQL db connection and
        add collected data to database.
        database ke table mei 5 columns rakh,
        4 for collected data, 1 for marking present.
        add if to check if email already in db,
        if yes give GUI prompt to confirms    
        """

    qrInput = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()
   
    global root, btn, bnt   
    root = Tk()
    root.title("Select")        
    root.geometry("285x100")
    root.resizable(False, False)
    root.config(background="green")    
    btn = Button(root, height='1', width=13, text="QR Generator", pady=5, command=QRGen).pack()
    Label(root, text="", height='1', width='0').pack()
    bnt = Button(root, height='1', width=13, text="QR Scanner", pady=5, command=QRScan).pack()
    
    """
    btn = Button(height='1', width=13, text="QR Generator", command=QRGen)
    bnt = Button(height='1', width=13, text="QR Scanner", command=QRScan)
    btn.grid(row=1, column=2, padx=5, pady=5, columnspan=1)
    bnt.grid(row=1, column=3, padx=5, pady=5, columnspan=1)
    root.mainloop()
    btn.pack()
    bnt.pack()
    """

def main_page():
    def register_user():
        username_info = username.get()
        password_info = password.get()
        file = open(username_info,"w")
        file.write(username_info+"\n")
        file.write(password_info)
        file.close()
        Label(screen1, text="Registeration Success", font ="green", height='2', width='30').pack()

    def register():
        global screen1
        screen1 = Toplevel(screen)
        screen1.geometry("350x250")
        screen1.title("Register")
        global username 
        global password
        global username_entry
        global password_entry
        username = StringVar()
        password = StringVar()
        Label(screen1, text="Please enter your information", height="2", width="30").pack()
        Label(screen1, text="", height='1', width='30').pack()
        Label(screen1, text="Username", height='1', width='30').pack()
        username_entry = Entry(screen1, textvariable=username)
        username_entry.pack()
        Label(screen1, text="Password", height='2', width='30').pack()
        password_entry = Entry(screen1, show="*", textvariable=password)
        password_entry.pack()
        Label(screen1, text="", height='1', width='30').pack()
        Button(screen1, text="Sumbit", height='1', width='18', command=register_user).pack()

    def login():
        global screen2
        global username_verify
        global password_verify
        username_verify = StringVar()
        password_verify = StringVar()
        screen2 = Toplevel(screen)
        screen2.geometry("300x250")
        screen2.title("Login")
        Label(screen2, text="Please enter your login information", height='2', width='30').pack()
        Label(screen2, text="",height='1', width='30').pack()
        Label(screen2, text="Username : ", height='1', width='30').pack()
        username_entry1 = Entry(screen2, textvariable=username_verify)
        username_entry1.pack()
        Label(screen2, text="Password : ", height='1', width='30').pack()
        password_entry1 = Entry(screen2, show="*", textvariable=password_verify)
        password_entry1.pack()
        Label(screen2, text="", height='1', width='30').pack()
        Button(screen2, text="Login", height="1", width="18", command=login_verify).pack()

    def login_verify():
        global screen3
        screen3 = Toplevel(screen)
        screen3.geometry("150x30")
        username1 = username_verify.get()
        password1 = password_verify.get()
        list_of_dir = os.listdir()
        if username1 in list_of_dir:
            file = open (username1, "r")
            verify = file.read().splitlines()
            if password1 in verify:
                screen3.title("Info")
                screen3.geometry("150x65")
                Label(screen3, text="Login Success", height='1', width='30', pady=5).pack()
                Button(screen3, text="OK", height="1", width="10", command=mgm_page, pady=5).pack()
            else:
                screen3.title("Alert")
                Label(screen3, text="Incorrect Password", height='1', width='30', pady=5).pack()

        else :
            screen3.title("Alert")
            Label(screen3, text="Invalid User", height='1', width='30', pady=5).pack()

    def acc_screen():
        global screen
        screen = Tk()
        screen.title("Event Management")
        screen.geometry("300x225")
        Label(text="", height='1', width='30').pack()
        Label(text="Sign Up if you are new User, \n Login if existing User.").pack()
        Label(text="", height='1', width='30').pack()
        Button(text="Login ", height='2', width='30', command=login).pack()
        Label(text="",height='1', width='30').pack()
        Button(text="Register", height='2', width='30', command=register).pack()
        screen.mainloop()
        
    acc_screen()

main_page()

